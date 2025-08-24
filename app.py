# app.py
import re, json, io, math, os, tempfile
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from trace_report import _kv_from_wo, _actions_from_wo, _details_from_json, build_details_df_from_payload
from _to_kml import to_kml  

# --------------------------- Helpers ---------------------------
# --- ALWAYS-WORK fallback writer (plain .xlsx, no macros) ---
def simple_workbook_bytes(meta: dict, wo_df: pd.DataFrame, details_df: pd.DataFrame) -> bytes:
    if details_df is None or not hasattr(details_df, "to_excel"):
        details_df = pd.DataFrame()
    buf = io.BytesIO()
    # close/flush via context manager (prevents 0B files)
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        pd.DataFrame([meta]).to_excel(writer, index=False, sheet_name="Meta")
        wo_df.to_excel(writer, index=False, sheet_name="WO")
        details_df.to_excel(writer, index=False, sheet_name="Details")
    buf.seek(0)
    return buf.getvalue()

# --- Normalize build_excel return into bytes; detect macro workbook ---
def normalize_excel_output(x, *, fallback_meta=None, fallback_wo=None, fallback_details=None):
    """
    Returns (bytes, ext, mime). Tries x first; if unusable, falls back to simple .xlsx.
    ext in {'.xlsx', '.xlsm'} with matching MIME.
    """
    # 1) If already bytes
    if isinstance(x, (bytes, bytearray)):
        return bytes(x), ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # 2) openpyxl workbook (with/without macros)
    try:
        from openpyxl.workbook import Workbook as _WB
        if isinstance(x, _WB):
            # If wb contains macros, openpyxl marks it via wb.vba_archive
            has_vba = getattr(x, "vba_archive", None) is not None
            buf = io.BytesIO()
            x.save(buf)        # IMPORTANT: save() writes and closes internal state
            buf.seek(0)
            return (
                buf.getvalue(),
                ".xlsm" if has_vba else ".xlsx",
                "application/vnd.ms-excel.sheet.macroEnabled.12" if has_vba
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception:
        pass

    # 3) Path on disk
    if isinstance(x, str) and os.path.exists(x):
        ext = os.path.splitext(x)[1].lower()
        mime = (
            "application/vnd.ms-excel.sheet.macroEnabled.12" if ext == ".xlsm"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        with open(x, "rb") as f:
            return f.read(), (".xlsm" if ext == ".xlsm" else ".xlsx"), mime

    # 4) DataFrame (rare, but handle)
    if isinstance(x, pd.DataFrame):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
            x.to_excel(w, index=False, sheet_name="Sheet1")
        buf.seek(0)
        return buf.getvalue(), ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # 5) Nothing usable → fallback to a minimal valid .xlsx
    if fallback_meta is not None and fallback_wo is not None:
        b = simple_workbook_bytes(fallback_meta, fallback_wo, fallback_details or pd.DataFrame())
        return b, ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # If we get here, signal error up the stack
    raise RuntimeError(f"Unsupported build_excel() return type: {type(x)}")

def require_inputs(up_wo_csv, up_json):
    """Validate uploads, read them safely, or stop with a user-facing warning."""
    if up_wo_csv is None or up_json is None:
        st.warning("Please upload the **WO CSV** and the **circuit JSON** first.")
        st.stop()

    # Read CSV
    try:
        up_wo_csv.seek(0)
        wo_df = pd.read_csv(up_wo_csv)
    except Exception as e:
        st.warning(f"Could not read the WO CSV: {e}")
        st.stop()

    # Read JSON
    try:
        up_json.seek(0)
        payload = json.load(up_json)
    except Exception as e:
        st.warning(f"Could not read the circuit JSON: {e}")
        st.stop()

    return wo_df, payload

def _coalesce(d: dict, keys, default=None):
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return default

def _try_col(df: pd.DataFrame, names):
    for n in names:
        if n in df.columns:
            return n
    return None

def count_breaks_and_splices(df: pd.DataFrame, desc_col: str):
    breaks = int(df[desc_col].astype(str).str.strip().str.startswith(("BREAK", "Break")).sum())
    splices = int(df[desc_col].astype(str).str.strip().str.startswith(("Splice", "splice")).sum())
    return breaks, splices

def parse_endpoints_from_json(payload: dict):
    """
    Best-effort: looks for A/Z end strings in common keys.
    """
    # Flatten once
    flat = {}
    def walk(x, prefix=""):
        if isinstance(x, dict):
            for k, v in x.items():
                walk(v, f"{prefix}.{k}" if prefix else k)
        elif isinstance(x, list):
            for i, v in enumerate(x):
                walk(v, f"{prefix}[{i}]")
        else:
            flat[prefix] = x
    walk(payload)

    # naive search
    a_end = None
    z_end = None
    for k, v in flat.items():
        ks = k.lower()
        if isinstance(v, (str, int, float)):
            vs = str(v).strip()
            if a_end is None and ("a end" in ks or "a_end" in ks or ks.endswith(".a")):
                a_end = vs
            if z_end is None and ("z end" in ks or "z_end" in ks or ks.endswith(".z")):
                z_end = vs
    return a_end, z_end

def build_excel(meta: dict, wo_df: pd.DataFrame, details_df: pd.DataFrame, payload_json: dict) -> bytes:
    """
    Creates the Excel with two sheets:
      - Summary (header block + action table)
      - Details (details_df + 'Map It' hyperlinks when lat/lon available)
    """

    if details_df is None or not hasattr(details_df, "copy"):
        details_df = pd.DataFrame()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        # --- write your sheets here ---
        pd.DataFrame([meta]).to_excel(writer, index=False, sheet_name="Meta")
        wo_df.to_excel(writer, index=False, sheet_name="WO")

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Styles
        hfill = PatternFill("solid", fgColor="D9D9D9")
        bold = Font(bold=True)
        thin = Side(border_style="thin", color="AAAAAA")
        box = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        # ---------- Header block ----------
        # Derive counts and lengths
        desc_col = _try_col(wo_df, ["Description", "description", "DESC", "Desc"]) or wo_df.columns[0]
        action_col = _try_col(wo_df, ["Action", "ACTION", "Wo Action#", "WO Action#"]) or wo_df.columns[0]

        num_breaks, num_splices = count_breaks_and_splices(wo_df, desc_col)

        # lengths best‑effort: if single big numbers exist, use them; else sum
        end_to_end_len = None
        otdr_len = None
        len_col = _try_col(wo_df, ["Length", "length", "End to End Length(m)", "End_to_End_Length_m"])
        otdr_col = _try_col(wo_df, ["~OTDR Length", "~OTDR Length(m)", "OTDR Length", "otdr_length"])

        if len_col:
            try:
                vals = pd.to_numeric(wo_df[len_col], errors="coerce").dropna()
                end_to_end_len = int(vals.sum()) if len(vals) > 1 else int(vals.iloc[0])
            except:
                end_to_end_len = None
        if otdr_col:
            try:
                vals = pd.to_numeric(wo_df[otdr_col], errors="coerce").dropna()
                otdr_len = int(vals.sum()) if len(vals) > 1 else int(vals.iloc[0])
            except:
                otdr_len = None

        # A/Z ends from JSON (fallback to meta)
        a_end_json, z_end_json = parse_endpoints_from_json(payload_json or {})
        a_end = a_end_json or meta.get("a_end")
        z_end = z_end_json or meta.get("z_end")

        # Header grid values (match your screenshot text)
        left_labels = [
            ("Order Number:", meta.get("order_id")),
            ("Work Order Number:", meta.get("wo_id")),
            ("Order A to Z:", f"{a_end}_{z_end}" if a_end and z_end else ""),
            ("Designer:", meta.get("designer_name")),
            ("Contact Number:", meta.get("designer_phone")),
            ("Date (dd/mm/yyyy):", meta.get("date")),
            ("Details:", meta.get("details")),
            ("ORDER Number:", meta.get("order_id")),
        ]
        right_labels = [
            ("Number of Fibre Breaks:", num_breaks),
            ("Number of Fibre Splices", num_splices),
            ("End to End Length(m)", end_to_end_len),
            ("End to End ~ OTDR(m)", otdr_len),
            ("A END:", a_end),
            ("Z END:", z_end),
            ("Work Order Processing Results", ""),
            ("# of WO Splice/Locations:", 0),
            ("# of ACTION Splice/Locations:", 0),
        ]

        # write grid (two columns of label/value repeated)
        r = 1
        for label, value in left_labels[:6]:
            ws.cell(row=r, column=1, value=label).font = bold
            ws.cell(row=r, column=2, value=value)
            if r <= 4:  # first 4 rows have a matching right-side line
                rl = right_labels[r-1]
                ws.cell(row=r, column=9, value=rl[0]).font = bold
                ws.cell(row=r, column=10, value=rl[1])
            r += 1
        # continue the remaining right labels beneath
        start_r = 5
        for i, (label, value) in enumerate(right_labels[4:], start=start_r):
            ws.cell(row=i, column=9, value=label).font = bold
            ws.cell(row=i, column=10, value=value)

        # Last two left rows
        ws.cell(row=7, column=1, value="Details:").font = bold
        ws.cell(row=7, column=2, value=meta.get("details") or "OSP DF")
        ws.cell(row=8, column=1, value="ORDER Number:").font = bold
        ws.cell(row=8, column=2, value=meta.get("order_id"))

        # ---------- Actions table ----------
        start_row = 10
        ws.cell(row=start_row, column=1, value="Action").font = bold
        ws.cell(row=start_row, column=2, value="Description").font = bold
        ws.cell(row=start_row, column=9, value="SAP").font = bold  # same columns as screenshot

        # bring in only the two columns we care about
        actions = wo_df[[c for c in [action_col, desc_col] if c in wo_df.columns]].copy()
        actions.columns = ["Action", "Description"]

        # Sort numerically if action has a number at start
        def sort_key(x):
            m = re.match(r"^\s*(\d+)", str(x))
            return int(m.group(1)) if m else 10**9
        actions = actions.sort_values(by="Action", key=lambda s: s.map(sort_key))

        for i, (_, row) in enumerate(actions.iterrows(), start=start_row+1):
            ws.cell(row=i, column=1, value=row["Action"])
            ws.cell(row=i, column=2, value=row["Description"])
            ws.cell(row=i, column=9, value="")  # SAP blank per screenshot

        # width + looks
        for col, width in [(1,18),(2,120),(9,12),(10,28)]:
            ws.column_dimensions[get_column_letter(col)].width = width
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
            for c in r:
                c.alignment = Alignment(vertical="center")
                if c.row == start_row:
                    c.fill = hfill
                    c.font = bold

        # -------------------- Details sheet --------------------
        ws2 = wb.create_sheet("Details")

        df = details_df.copy()

        # Add Map It hyperlink if we can find lat/lon
        lat_col = _try_col(df, ["lat", "Lat", "latitude", "Latitude", "LAT"])
        lon_col = _try_col(df, ["lon", "Lon", "lng", "Lng", "longitude", "Longitude", "LON"])

        if lat_col and lon_col:
            maps_col = "Map It"
            df[maps_col] = ""
        else:
            maps_col = None

        # Write header
        for j, col in enumerate(df.columns, start=1):
            ws2.cell(row=1, column=j, value=col).font = bold
            ws2.cell(row=1, column=j).fill = hfill
            ws2.cell(row=1, column=j).alignment = center

        # Write rows
        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws2.cell(row=i, column=j, value=val)

        # Insert hyperlinks
        if maps_col:
            mc_index = df.columns.get_loc(maps_col) + 1
            lat_idx = df.columns.get_loc(lat_col) + 1
            lon_idx = df.columns.get_loc(lon_col) + 1
            for i in range(2, 2 + len(df)):
                lat_cell = f"{get_column_letter(lat_idx)}{i}"
                lon_cell = f"{get_column_letter(lon_idx)}{i}"
                link = f'=HYPERLINK("https://www.google.com/maps?q="&{lat_cell}&","&{lon_cell},"Google Maps")'
                ws2.cell(row=i, column=mc_index, value=link)

        # basic widths
        for j in range(1, ws2.max_column + 1):
            ws2.column_dimensions[get_column_letter(j)].width = 24

        # Return bytes
        buf = BytesIO()
        wb.save(buf)
        return buf.read()
    
    buf.seek(0)
    return buf.getvalue()       

def read_uploaded_table(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        raise ValueError("No file uploaded")
    name = (uploaded_file.name or "").lower()
    data = uploaded_file.getvalue()  # bytes
    bio = io.BytesIO(data)

    if name.endswith((".xlsx", ".xlsm", ".xls")):
        bio.seek(0)
        return pd.read_excel(bio, engine="openpyxl")

    # CSV: autodetect delimiter + tolerant encodings
    for enc in ("utf-8", "utf-16", "latin1"):
        try:
            bio.seek(0)
            return pd.read_csv(
                bio, sep=None, engine="python", on_bad_lines="skip",
                encoding=enc, encoding_errors="ignore"
            )
        except Exception:
            continue

    bio.seek(0)
    return pd.read_excel(bio, engine="openpyxl")  # last resort

def _guess_latlon_cols(df):
    cols = {c.lower(): c for c in df.columns}
    lat = next((cols[c] for c in ["lat","latitude","y","lat_dd"] if c in cols), None)
    lon = next((cols[c] for c in ["lon","lng","longitude","x","lon_dd"] if c in cols), None)
    return lat, lon

def placemarks_from_wo_df(wo_df):
    lat_col, lon_col = _guess_latlon_cols(wo_df)
    if not lat_col or not lon_col:
        return []
    # choose a name-ish column if available
    name_col = None
    for c in wo_df.columns:
        cl = c.lower()
        if any(k in cl for k in ["name","site","id","device","label","address","location"]):
            name_col = c; break

    pms = []
    for _, r in wo_df.iterrows():
        try:
            lat, lon = float(r[lat_col]), float(r[lon_col])
        except Exception:
            continue
        if not (math.isfinite(lat) and math.isfinite(lon)):
            continue
        name = str(r[name_col]).strip() if name_col and pd.notna(r.get(name_col)) else "WO Point"
        pms.append({"name": name, "lat": lat, "lon": lon, "description": ""})
    return pms

# parses lines like: "... Address: ... () : 43.644719, -79.385046 :  : something"
_COORD_RE = re.compile(r":\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*:\s*:\s*(.*)$")
def placemarks_from_payload(payload):
    # Walk JSON to collect candidate strings
    texts=[]
    def walk(x):
        if isinstance(x, dict):
            for v in x.values(): walk(v)
        elif isinstance(x, list):
            for v in x: walk(v)
        elif isinstance(x, str):
            if "Address:" in x and ":" in x:
                texts.append(x)
    walk(payload)

    pms=[]
    for blob in texts:
        for line in blob.splitlines():
            m = _COORD_RE.search(line)
            if not m: 
                continue
            lat = float(m.group(1)); lon = float(m.group(2))
            tail = m.group(3).strip()
            # try to derive a short name from the part before "Address"
            pre = line.split("Address",1)[0]
            parts = [p.strip() for p in pre.split(",") if p.strip()]
            name = parts[1] if len(parts)>=2 else "JSON Point"
            desc = tail
            pms.append({"name": name, "lat": lat, "lon": lon, "description": desc})
    return pms

def dedupe_placemarks(placemarks):
    seen=set(); out=[]
    for pm in placemarks:
        key=(round(pm.get("lat",0),6), round(pm.get("lon",0),6), pm.get("name","").upper())
        if key in seen: 
            continue
        seen.add(key); out.append(pm)
    return out

# --------------------------- Streamlit UI ---------------------------
ROUTE_TYPES   = ["Primary", "Diverse", "Triverse"]
CIRCUIT_TYPES = ["NEW", "Existing"]
st.set_page_config(page_title="Trace Builder", layout="wide")
st.title("Trace Builder (CSV + JSON → Excel / KML)")
st.caption("Upload WO.csv + WO.json, choose options, then export the Excel report (Cover Page, Fibre Trace, Activity Overview Map, Error Report) and a colorized KML/KMZ.")

# --sidebar inputs
st.sidebar.header("Project metadata")
st.sidebar.markdown("Fill in the project metadata to generate a complete workbook.")
with st.sidebar:
    st.header("Project metadata")
    designer_name  = st.text_input("Designer Name", "abcd")
    designer_email = st.text_input("Designer Email", "abcd@beanfield.com")
    designer_phone = st.text_input("Designer Phone #", "4164164164")
    fibers_assigned = st.number_input("# of Fibers Assigned", min_value=1, step=1, value=2)

    order_number   = st.text_input("Order Number / Circuit ID", "ORDER-267175")
    circuit_id     = st.text_input("Circuit ID", "CK37105")
    client_name    = st.text_input("Client Name", "ABC In")
    service_type   = st.text_input("Service Type", "IRU Fbr")
    device_type    = st.text_input("Device Type", "N/A")

    route_type     = st.selectbox("Route Type", ROUTE_TYPES, index=0)
    circuit_type   = st.selectbox("Select Circuit Type", CIRCUIT_TYPES, index=0)

    a_end          = st.text_input("A End (auto)", "BFMH-0021")
    z_end          = st.text_input("Z End (auto)", "SITE5761")
    circuit_version= st.text_input("Circuit Version", "1")

    template_xlsm  = st.file_uploader("Macro template (.xlsm, optional)", type=["xlsm"])
    want_xlsm      = st.checkbox("Export as .xlsm (requires template)", value=bool(template_xlsm))


st.divider()
# --------------------------------- UPLOADS ----------------------------------
st.subheader("0) Before Upload Files...")
st.markdown("Create your circuit as per normal, and save the Fibre Trace as a JSON File")
st.subheader("1) Upload inputs")
col1, col2 = st.columns(2)


with col1:
    with st.container(border=True):
        st.markdown("Upload Work Order CSV file")
        up_wo_csv = st.file_uploader("WO.csv", type=["csv"], label_visibility="collapsed")
with col2: 
    with st.container(border=True):
        st.markdown("Upload Work Order JSON file")
        up_json   = st.file_uploader("WO.json", type=["json"],label_visibility="collapsed")

ready = (up_wo_csv is not None) and (up_json is not None)
if not ready:
    st.info("Upload both files to enable generation.")
else:
    wo_df = read_uploaded_table(up_wo_csv)
    # use functions from this script:
    wo_kv = _kv_from_wo(wo_df)
    actions_df = _actions_from_wo(wo_df)
    details_df = _details_from_json(up_json)

    # Assume you already have:
# up_wo_csv = st.file_uploader("Upload WO CSV", type=["csv"], key="wo_csv")
# up_json   = st.file_uploader("Upload circuit JSON", type=["json"], key="circ_json")




# up_wo_csv.seek(0) # Reset buffer for reading
# up_json.seek(0) # Reset buffer for reading


# --------------------------------- Generate ----------------------------------
st.subheader("2) Generate outputs")

# --- Generate workbook & KML side by side ---
col_gen1, col_gen2 = st.columns(2)

# ---------- Remove & Add: CSV preview + Generate Excel ----------
from remove_add_algo import transform_remove_add
with st.container(border=True):
    st.markdown("### Remove & Add — CSV Generator")

    ra_file = st.file_uploader("Upload Remove & Add input (CSV or XLSX)", type=["csv","xlsx"], key="ra_input")

    if st.button("Generate", type="primary", key="btn_ra_generate", disabled=(ra_file is None)):
        try:
            ra_file.seek(0)
            if ra_file.name.lower().endswith(".csv"):
                df_in = pd.read_csv(ra_file)
            else:
                df_in = pd.read_excel(ra_file)

            df_out = transform_remove_add(df_in)

            st.caption("Preview (first 200 rows)")
            st.dataframe(df_out.head(200), use_container_width=True, hide_index=True)

            st.download_button(
                "Download Remove & Add.csv",
                df_out.to_csv(index=False, encoding="utf-8-sig"),
                file_name="Remove & Add.csv",
                mime="text/csv",
                key="dl_remove_add_csv",
            )
        except Exception as e:
            st.error(f"Failed to generate CSV: {e}")


with col_gen1:
    # --------------------- Activity Overview Map ---------------------
    from parse_device_sheet import main as parse_device_main  # for activity overview parsing
    with st.container(border=True):
        st.markdown("Activity Overview Map")
        # The button that triggers the generation
        if st.button("Generate", type="primary", key="btn_activity_overview", disabled=not ready):
            # Make sure inputs are present and readable
            wo_df, payload = require_inputs(up_wo_csv, up_json)

            # Try to use the exact logic from parse_device_sheet.py
            try:
                from parse_device_sheet import gather_connections, parse_device_table  # exact logic

            except Exception:
                gather_connections = None
                parse_device_table = None

            # Fallback gatherer if functions weren't exported in the module
            def _fallback_gather_connections(obj):
                out = []
                def rec(x):
                    if isinstance(x, dict):
                        for k, v in x.items():
                            if k == "Connections" and isinstance(v, str):
                                out.append(v)
                            else:
                                rec(v)
                    elif isinstance(x, list):
                        for it in x:
                            rec(it)
                rec(obj)
                return out

            # Decide which gatherer we’ll use
            _gather = gather_connections or _fallback_gather_connections

            try:
                conns = _gather(payload)
                if not conns:
                    st.error("No 'Connections' strings found in the JSON under 'Report: Splice Details'.")
                    st.stop()

                # Use the first unique string (how the script behaves)
                conn_text = conns[0]

                # If parse_device_table is available, use it; otherwise fail loudly (you want exact logic)
                if parse_device_table is None:
                    raise RuntimeError("`parse_device_table` not found in parse_device_sheet.py. Please export it.")

                # Build the table (append_details=True keeps richer rows if your script supports it)
                df = parse_device_table(conn_text, append_details=True)

                if df is None or df.empty:
                    st.warning("Activity Overview Map produced an empty table.")
                    st.stop()

                # Prepare CSV bytes and a preview
                csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

                st.success(f"Generated Activity Overview table with {len(df)} rows.")
                st.download_button(
                    "Download Activity Overview CSV",
                    data=csv_bytes,
                    file_name="activity_overview.csv",
                    mime="text/csv",
                    key="dl_activity_overview_csv",
                )

                # Preview (responsive, no index)
                st.dataframe(df, use_container_width=True, hide_index=True)

            except Exception as e:
                st.error(f"Failed to generate Activity Overview Map: {e}")

    # --------------------- Fiber Trace Button ---------------------
     # Import the helper above (make sure fibre_trace.py is alongside app.py or in PYTHONPATH)
    from fibre_trace import generate_xlsx  # generate(json_path, out_csv=None, out_xlsx=None, out_kml=None, ug=100, ar=0)
    # --- inside your UI layout ---
    with st.container(border=True):
        st.markdown("Fibre Trace")

        if st.button("Generate", type="primary", key="btn_fibre_trace", disabled=not ready):
            try:
                # Save uploaded JSON to a temp file
                up_json.seek(0)
                with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as tmp:
                    tmp.write(up_json.read())
                    json_path = Path(tmp.name)

                # Output path (temp)
                out_xlsx = json_path.with_name("fibre_trace.xlsx")

                # Generate Excel only (no CSV/KML)
                df_trace = generate_xlsx(json_path, out_xlsx)

                # Download Excel
                with open(out_xlsx, "rb") as fh:
                    st.download_button(
                        "Download Fibre Trace (Excel)",
                        fh,
                        file_name="fibre_trace.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_fibre_trace_xlsx",
                    )

                # Preview (optional)
                st.success(f"Fibre Trace generated: {len(df_trace)} rows.")
                st.dataframe(df_trace, use_container_width=True, hide_index=True)

            except Exception as e:
                st.error(f"Failed to generate Fibre Trace: {e}")



with col_gen2:
    # --------------------- Generate KML Button ---------------------
    with st.container(border=True):
        st.markdown ("KML")
        if st.button("Generate", type="primary", key="btn_generate_kml", disabled=not ready):
            if not up_wo_csv or not up_json:
                st.error("Please provide both WO.csv and the circuit JSON.")
                st.stop()

            wo_df = pd.read_csv(up_wo_csv)
            payload = json.load(up_json)

            pms = placemarks_from_wo_df(wo_df) + placemarks_from_payload(payload)
            pms = dedupe_placemarks(pms)

            kml_str = to_kml(title=order_number or "Trace", placemarks=pms)  # ✅ correct signature
            kml_bytes = kml_str.encode("utf-8")

            st.success(f"KML created with {len(pms)} placemarks.")
            st.download_button(
                "Download KML",
                data=kml_bytes,
                file_name=f"{order_number}_Trace.kml",
                mime="application/vnd.google-earth.kml+xml",
                key="dl_kml",
            )

    # --------------------- Genrate   ---------------------         