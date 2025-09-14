# fiber_action.py
from __future__ import annotations

import io
import json
import re
from typing import Dict, IO, List, Optional, Set, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# Regex helpers
# =========================
PAIR_RE = re.compile(
    r'(?P<verb>Splice|BREAK|Remove\s+splicing)\s+'
    r'(?P<A>[A-Za-z0-9#._/-]+)\s*\[\s*(?P<Ar>[^\]]+?)\s*\]\s*'
    r'(?:-|\s+)?\s*'
    r'(?P<B>[A-Za-z0-9#._/-]+)\s*\[\s*(?P<Br>[^\]]+?)\s*\]',
    re.IGNORECASE,
)

BOX_HEADER_RE = re.compile(r'^\s*([A-Za-z0-9#._/-]{2,})\s*:\s*OSP Splice Box\b.*$', re.IGNORECASE)
CA_PMD_RE = re.compile(r'^\s*CA\d+.*?PMID[^\d]*([0-9A-Z]+)\b', re.IGNORECASE)
PAIR_ORIENT_RE = re.compile(
    r'^\s*PMID[^0-9A-Z]*([0-9A-Z]+)\b.*?--\s*Splice\s*--\s*PMID[^0-9A-Z]*([0-9A-Z]+)\b',
    re.IGNORECASE,
)

TRAILING_MONTREAL_RE = re.compile(r'\s*-\s*Montreal\s*\(.*?\)\s*$', re.IGNORECASE)
EQUIP_PREFIX_RE     = re.compile(r'^\(equipment\b.*?\)\.?\s*', re.IGNORECASE)

DASH_NORMALIZER = dict.fromkeys(map(ord, "\u2010\u2011\u2012\u2013\u2014\u2212"), ord("-"))

# =========================
# CSV reading
# =========================
def _read_actions_csv(upload: Union[str, IO[bytes], bytes]) -> pd.DataFrame:
    try:
        if isinstance(upload, bytes):
            text = upload.decode("utf-8", "ignore")
        elif hasattr(upload, "read"):
            upload.seek(0)
            text = upload.read().decode("utf-8", "ignore")
        else:
            text = open(upload, "r", encoding="utf-8").read()
    except Exception:
        return pd.DataFrame(columns=["Action", "Description", "SAP"])

    lines = text.splitlines()
    hdr_idx = None
    for i, line in enumerate(lines):
        if line.strip().lower().startswith("action,description"):
            hdr_idx = i
            break

    try:
        if hdr_idx is not None:
            df = pd.read_csv(io.StringIO("\n".join(lines[hdr_idx:])),
                             dtype=str, keep_default_na=False)
        else:
            df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False)
    except Exception:
        return pd.DataFrame(columns=["Action", "Description", "SAP"])

    rename = {}
    for c in list(df.columns):
        lc = c.strip().lower()
        if lc == "action": rename[c] = "Action"
        if lc == "description": rename[c] = "Description"
        if lc == "sap": rename[c] = "SAP"
    df = df.rename(columns=rename)

    for col in ("Action", "Description"):
        if col not in df.columns:
            df[col] = ""
    if "SAP" not in df.columns:
        df["SAP"] = ""

    return df[["Action", "Description", "SAP"]].copy()

# =========================
# JSON → CA order maps
# =========================
def _extract_connection_blobs(json_source: Optional[Union[str, IO[bytes], bytes]]) -> List[str]:
    if json_source is None:
        return []
    try:
        if isinstance(json_source, bytes):
            raw = json_source.decode("utf-8", "ignore")
        elif hasattr(json_source, "read"):
            json_source.seek(0)
            raw = json_source.read().decode("utf-8", "ignore")
        else:
            raw = open(json_source, "r", encoding="utf-8").read()
        data = json.loads(raw)
    except Exception:
        return []
    blobs: List[str] = []
    try:
        for blk in data.get("Report: Splice Details", [])[0].get("", []):
            s = blk.get("Connections")
            if isinstance(s, str) and s.strip():
                blobs.append(s)
    except Exception:
        pass
    return blobs

def _parse_json_for_order(json_source: Optional[Union[str, IO[bytes], bytes]]
                          ) -> Tuple[Dict[str, Dict[str, int]], Dict[str, Set[str]], Dict[frozenset, Tuple[str, str]]]:
    ca_order_by_box: Dict[str, Dict[str, int]] = {}
    boxes_by_pmid: Dict[str, Set[str]] = {}
    pair_orient: Dict[frozenset, Tuple[str, str]] = {}

    for blob in _extract_connection_blobs(json_source):
        current_box: Optional[str] = None
        order_map: Dict[str, int] = {}
        for raw_line in blob.splitlines():
            line = raw_line.replace("<COLON>", ":").replace("<COMMA>", ",")
            m = BOX_HEADER_RE.match(line)
            if m:
                if current_box and order_map and current_box not in ca_order_by_box:
                    ca_order_by_box[current_box] = dict(order_map)
                current_box = m.group(1).strip()
                order_map = {}
                continue
            m = CA_PMD_RE.match(line)
            if m and current_box:
                pmid = m.group(1).strip()
                if pmid not in order_map:
                    order_map[pmid] = len(order_map)
                    boxes_by_pmid.setdefault(pmid, set()).add(current_box)
                continue
            m = PAIR_ORIENT_RE.match(line)
            if m:
                a, b = m.group(1).strip(), m.group(2).strip()
                pair_orient[frozenset((a, b))] = (a, b)
        if current_box and order_map and current_box not in ca_order_by_box:
            ca_order_by_box[current_box] = dict(order_map)

    return ca_order_by_box, boxes_by_pmid, pair_orient

# =========================
# Normalization helpers
# =========================
def _compact_range(r: str) -> str:
    r = str(r or "").translate(DASH_NORMALIZER)
    r = re.sub(r"\s*-\s*", "-", r)
    return r.strip()

def _clean_tail_after(source: str, match_end: int) -> str:
    tail = source[match_end:]
    tail = tail.lstrip(" -.,;")
    tail = EQUIP_PREFIX_RE.sub("", tail)
    tail = TRAILING_MONTREAL_RE.sub("", tail)
    tail = re.sub(r"\s+", " ", tail).strip()
    return tail

def _choose_box_by_pmid_order(
    A: str, B: str,
    ca_order_by_box: Dict[str, Dict[str, int]],
    boxes_by_pmid: Dict[str, Set[str]],
) -> Optional[str]:
    common = boxes_by_pmid.get(A, set()) & boxes_by_pmid.get(B, set())
    if not common:
        return None
    if len(common) == 1:
        return next(iter(common))
    for bx in common:
        om = ca_order_by_box.get(bx, {})
        if A in om and B in om and om[A] <= om[B]:
            return bx
    return sorted(common)[0]

def _normalize_desc(action: str, desc: str,
                    ca_order_by_box: Dict[str, Dict[str, int]],
                    boxes_by_pmid: Dict[str, Set[str]],
                    pair_orient: Dict[frozenset, Tuple[str, str]]) -> str:
    s = str(desc or "")
    s = s.translate(DASH_NORMALIZER)
    s = re.sub(r"\s+", " ", s).strip()

    m = PAIR_RE.search(s)
    if not m:
        return s

    if "remove" in (action or "").lower() or m.group("verb").lower().startswith("remove"):
        verb = "BREAK"
    elif "break" in (action or "").lower():
        verb = "BREAK"
    else:
        verb = "Splice"

    A, B = m.group("A"), m.group("B")
    Ar, Br = _compact_range(m.group("Ar")), _compact_range(m.group("Br"))

    box = _choose_box_by_pmid_order(A, B, ca_order_by_box, boxes_by_pmid)
    if box:
        om = ca_order_by_box.get(box, {})
        if A in om and B in om and om[A] > om[B]:
            A, B, Ar, Br = B, A, Br, Ar
    else:
        key = frozenset((A, B))
        if key in pair_orient:
            a, b = pair_orient[key]
            if {A, B} == {a, b} and A == b:
                A, B, Ar, Br = B, A, Br, Ar

    core = f"{verb} {A} [{Ar}] {B} [{Br}]"
    tail = _clean_tail_after(s, m.end())
    if tail:
        core = f"{core} {tail}"
    return core

# ---------- color helpers (fix: don't use startswith; Action has indices like '0: Add') ----------
BLUE_GREEN_XLSX = "B2DFDB"  # light blue-green fill for Add/Splice
PURPLE_XLSX     = "E1BEE7"  # light purple fill for Remove/BREAK

def _is_add(action: str) -> bool:
    a = (action or "").lower()
    return "add" in a

def _is_remove(action: str) -> bool:
    a = (action or "").lower()
    return ("remove" in a) or ("break" in a)

def _assign_excel_fill(action: str, description: str) -> str:
    if _is_remove(action) or str(description).upper().startswith("BREAK"):
        return PURPLE_XLSX
    if _is_add(action) or str(description).upper().startswith("SPLICE"):
        return BLUE_GREEN_XLSX
    return "FFFFFF"  # safe default

# =========================
# Public API
# =========================
def compute_fibre_actions(
    csv_source: Union[str, IO[bytes], bytes],
    json_source: Optional[Union[str, IO[bytes], bytes]] = None,
    **_compat_kwargs,
) -> pd.DataFrame:
    try:
        df = _read_actions_csv(csv_source)
        if not isinstance(df, pd.DataFrame) or df.empty:
            return pd.DataFrame(columns=["Action", "Description", "SAP", "Color"])

        ca_order_by_box, boxes_by_pmid, pair_orient = _parse_json_for_order(json_source)

        mask = df["Description"].fillna("").str.contains(r"\b(Splice|BREAK|Remove)\b", case=False, regex=True)
        df = df.loc[mask].reset_index(drop=True)

        df["Description"] = [
            _normalize_desc(a, d, ca_order_by_box, boxes_by_pmid, pair_orient)
            for a, d in zip(df["Action"], df["Description"])
        ]
        df["Color"] = [_assign_excel_fill(a, d) for a, d in zip(df["Action"], df["Description"])]

        for col in ("Action", "Description", "SAP", "Color"):
            if col not in df.columns:
                df[col] = ""
        return df[["Action", "Description", "SAP", "Color"]]
    except Exception:
        return pd.DataFrame(columns=["Action", "Description", "SAP", "Color"])

# =========================
# Excel export (color rows)
# =========================
def actions_to_workbook_bytes(df: pd.DataFrame, *, sheet_name: str = "Fibre Action") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ["Action", "Description", "SAP"]
    ws.append(headers)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)
    box = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = box
        cell.fill = header_fill

    for _, r in df.iterrows():
        ws.append([r.get("Action", ""), r.get("Description", ""), r.get("SAP", "")])
        row_idx = ws.max_row
        fill_clr = (r.get("Color") or "FFFFFF").upper().replace("#", "")
        row_fill = PatternFill("solid", fgColor=fill_clr)
        for cell in ws[row_idx]:
            cell.border = box
            cell.alignment = wrap
            cell.fill = row_fill

    # simple auto-width
    for i, col in enumerate(headers, start=1):
        series = [str(col)] + [str(v) for v in df[col]]
        width = min(max(len(x) for x in series) + 2, 100)
        ws.column_dimensions[get_column_letter(i)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
