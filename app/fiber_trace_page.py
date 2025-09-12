"""
fiber_trace.py
Build the Fiber Trace workbook from ONE JSON ("Report: Splice Details") and ONE CSV (actions).

Outputs a single Excel sheet with:
- Equipment Location (dark green, only col A filled)
- Equipment (dark green)
- Cable[attach] (light green)
- Existing Splice (yellow, collapsed into ranges; # of Fbrs filled)
- Break Splice (red) for Remove(E) actions
- Splice Required (orange) for Add actions
- CableInfo (blue) only when length info exists
- Google Maps link placed on Equipment Location / Equipment rows

Public entrypoint: build_fiber_trace(json_src, csv_src, out_path) -> (out_path, stats)
"""

from __future__ import annotations
import re
import json
import collections
from typing import Any, Dict, Iterable, List, Tuple, Union, IO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ----------------------------- styling (fill ONLY column A) -----------------------------
FILL_DARK_GREEN = PatternFill("solid", fgColor="006400")
FILL_LIGHT_GREEN = PatternFill("solid", fgColor="A9D08E")
FILL_RED        = PatternFill("solid", fgColor="FF0000")
FILL_ORANGE     = PatternFill("solid", fgColor="F4B183")
FILL_BLUE       = PatternFill("solid", fgColor="9DC3E6")
FILL_YELLOW     = PatternFill("solid", fgColor="FFE699")


# --------------------------------------- helpers ---------------------------------------
def _clean(t: str) -> str:
    """Normalize placeholders and whitespace from the JSON export."""
    return (
        (t or "")
        .replace("<COLON>", ":")
        .replace("<OPEN>", "(")
        .replace("<CLOSE>", ")")
        .replace("<COMMA>", ",")
        .strip()
    )


def _split_sections(big: str) -> List[str]:
    """The JSON pack strings sections separated by '\n.\n' and some start with a leading '.'."""
    parts = [p.strip() for p in big.split("\n.\n") if p.strip()]
    return [p[1:] if p.startswith(".") else p for p in parts]


def _collapse(nums: Iterable[int]) -> List[Tuple[int, int]]:
    """Collapse sorted integers into contiguous ranges."""
    nums = sorted(set(nums))
    if not nums:
        return []
    out: List[Tuple[int, int]] = []
    s = prev = nums[0]
    for x in nums[1:]:
        if x == prev + 1:
            prev = x
        else:
            out.append((s, prev))
            s = prev = x
    out.append((s, prev))
    return out


# ------------------------------------ JSON parsing -------------------------------------
def _parse_section(section_text: str) -> Dict[str, Any]:
    lines = section_text.strip().splitlines()
    header = lines[0].strip()
    equipline = lines[2].strip() if len(lines) > 2 else ""
    cables, existing, cinfo = [], [], []
    for ln in lines[3:]:
        t = ln.strip()
        if not t:
            continue
        if t.startswith("CA") or "PMID<COLON>" in t:
            cables.append(t)
        elif "-- Splice --" in t:
            existing.append(t)
        elif "Cable Length Calculated" in t:
            cinfo.append(t)
    return {"header": header, "equipline": equipline, "cables": cables, "cinfo": cinfo, "existing": existing}


def _build_sections(json_src: Union[str, IO[str]]) -> List[Dict[str, Any]]:
    """Return list of equipment blocks (kept in A→Z order as in the JSON)."""
    if hasattr(json_src, "read"):
        data = json.load(json_src)  # type: ignore[arg-type]
    else:
        with open(json_src, "r", encoding="utf-8") as f:
            data = json.load(f)

    blocks = data["Report: Splice Details"][0][""]
    sections: List[Dict[str, Any]] = []

    for blk in blocks:
        for sec in _split_sections(blk["Connections"]):
            p = _parse_section(sec)

            # Header: "<City>,<Site>, Address..., : <lat>, <lng> : : <facility>"
            m = re.search(
                r"^(?P<city>[^,]+),(?P<site>[^,]+), Address.*?:\s*(?P<lat>-?\d+\.\d+),\s*(?P<lng>-?\d+\.\d+)\s*:\s*:\s*(?P<facility>.+)",
                p["header"],
            )
            if not m:
                continue

            city, site, lat, lng, facility = m.group("city", "site", "lat", "lng", "facility")
            eq_location = f"{city},{site},{facility}"
            eq_type = _clean(p["equipline"])

            # Cables & pmid→apt mapping
            pmid_to_apt: Dict[str, str] = {}
            cable_rows: List[str] = []
            for ca in p["cables"]:
                ct = _clean(ca)
                mpm = re.search(r"PMID[:\s]\s*(\w+)", ct)
                mapt = re.search(r"\(APT[:\s]*[: ]?([^)]+)\)", ct)
                if mpm:
                    pmid_to_apt[mpm.group(1)] = (mapt.group(1).strip() if mapt else "")
                cable_rows.append(ct)

            # Existing splices (collapse to ranges & set # of Fbrs)
            pair_to_fibA: Dict[Tuple[str, str], set[int]] = collections.defaultdict(set)
            for ex in p["existing"]:
                ex_t = _clean(ex)
                m1 = re.search(
                    r"PMID:\s*(\w+).*?F\[?(\d+)(?:\s*-\s*(\d+))?\]?\s*--\s*Splice\s*--\s*PMID:\s*(\w+).*?F\[?(\d+)(?:\s*-\s*(\d+))?\]?",
                    ex_t,
                )
                if m1:
                    a_p, a1, a2, b_p, b1, b2 = m1.groups()
                    rngA = range(int(a1), int(a2) + 1) if a2 else [int(a1)]
                    for v in rngA:
                        pair_to_fibA[(a_p, b_p)].add(v)
                else:
                    # F243 (no brackets) variant
                    m2 = re.search(r"PMID:\s*(\w+).*?F(\d+)\s*--\s*Splice\s*--\s*PMID:\s*(\w+).*?F(\d+)", ex_t)
                    if m2:
                        a_p, f1, b_p, _f2 = m2.groups()
                        pair_to_fibA[(a_p, b_p)].add(int(f1))

            existing_rows: List[Tuple[str, int]] = []
            for (pa, pb), fibs in pair_to_fibA.items():
                for a, b in _collapse(sorted(fibs)):
                    cnt = b - a + 1
                    aptA = pmid_to_apt.get(pa, "")
                    aptB = pmid_to_apt.get(pb, "")
                    desc = f"PMID: {pa}, Aptum ID: {aptA} .F[{a}-{b}] -- Existing Splice(s) -- PMID: {pb}, Aptum ID: {aptB} .F[{a}-{b}]"
                    existing_rows.append((desc, cnt))

            sections.append(
                {
                    "site_id": site.strip(),
                    "eq_location": eq_location,
                    "eq_type": eq_type,
                    "lat": float(lat),
                    "lng": float(lng),
                    "cables": cable_rows,
                    "pmid_to_apt": pmid_to_apt,
                    "existing": existing_rows,
                    "cinfo": [_clean(x) for x in p["cinfo"]],
                }
            )

    return sections


def _build_indexes(sections: List[Dict[str, Any]]):
    """Support routing actions to the correct equipment block."""
    # pmid -> set(site_ids)
    pmid_to_sites: Dict[str, set[str]] = {}
    # equipment id (e.g., M#1189A, D#30A) -> site id (e.g., PA27014)
    eq_to_site: Dict[str, str] = {}

    for s in sections:
        # pmid index
        for ca in s["cables"]:
            m = re.search(r"PMID[:\s]\s*(\w+)", ca)
            if m:
                pmid_to_sites.setdefault(m.group(1), set()).add(s["site_id"])

        # equipment-to-site map (parse first token of eq_type, e.g. "D#101 : OSP Splice Box - ...")
        m_eq = re.match(r"\s*([MD]#\w+|BFMA\d+)\s*:", s["eq_type"])
        if m_eq:
            eq_to_site[m_eq.group(1).replace(" ", "")] = s["site_id"]

    return pmid_to_sites, eq_to_site


# ------------------------------------ Actions parsing -----------------------------------
def _assign_action_site(desc: str, pmid_a: str, pmid_b: str,
                        pmid_to_sites: Dict[str, set[str]], eq_to_site: Dict[str, str]) -> str | None:
    # 1) explicit site/equipment hint inside description
    m = re.search(r"(PA ?\d{4,6}|M#[A-Za-z0-9]+|D#[A-Za-z0-9\-]+|BFMA\d+|BFAS-\d+)", desc)
    if m:
        token = m.group(1).replace(" ", "")
        return eq_to_site.get(token, token)  # map equipment to its site if we know it

    # 2) intersection of pmid locations
    sA = pmid_to_sites.get(pmid_a, set())
    sB = pmid_to_sites.get(pmid_b, set())
    inter = sA & sB
    if inter:
        return sorted(inter)[0]

    # 3) fallback to union
    uni = sA | sB
    return sorted(uni)[0] if uni else None


def _parse_actions(csv_src: Union[str, IO[str]], pmid_to_sites, eq_to_site):
    """Return list of normalized actions with placement site_id and fiber counts."""
    df = pd.read_csv(csv_src)
    actions: List[Dict[str, Any]] = []
    for _, r in df.iterrows():
        action = str(r.get("Action", ""))
        desc = str(r.get("Description", ""))

        # classify
        atype = "Splice Required" if ("Add" in action and "Remove" not in action) else "Break Splice"

        # fibers
        m = re.search(r"(?:Splice|BREAK)\s+(\w+)\s*\[(\d+)\s*-\s*(\d+)\]\s+(\w+)\s*\[(\d+)\s*-\s*(\d+)\]", desc)
        m_single = re.search(r"(?:Splice|BREAK)\s+(\w+)\s*\[(\d+)\]\s+(\w+)\s*\[(\d+)\]", desc)
        if m:
            ap, a1, a2, bp, b1, b2 = m.groups()
            a1, a2, b1, b2 = map(int, (a1, a2, b1, b2))
        elif m_single:
            ap, a1, bp, b1 = m_single.groups()
            a1 = a2 = int(a1)
            b1 = b2 = int(b1)
        else:
            # Not a splice action, skip (e.g., add equipment, connect ports, etc.)
            continue

        site = _assign_action_site(desc, ap, bp, pmid_to_sites, eq_to_site)

        actions.append(
            {
                "atype": atype,
                "site_id": site,
                "pmid_a": ap,
                "pmid_b": bp,
                "a1": a1,
                "a2": a2,
                "b1": b1,
                "b2": b2,
                "count": a2 - a1 + 1,
                "action_ref": action,
                "raw_desc": desc,
            }
        )
    return actions


# ------------------------------------ Workbook writer -----------------------------------
def _write_workbook(
    sections: List[Dict[str, Any]],
    actions: List[Dict[str, Any]],
    out_path: str,
) -> Dict[str, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fiber Trace"

    headers = ["Type", "Description", "# of Fbrs", "Action Ref", "Eq Location", "Eq Type", "Google Maps"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    def add_row(values: List[Any], fill: PatternFill | None = None):
        ws.append(values)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    # index actions by site
    site_to_actions: Dict[str, List[Dict[str, Any]]] = collections.defaultdict(list)
    for a in actions:
        site_to_actions[a["site_id"]].append(a)

    placed, unplaced = 0, 0

    # Emit blocks in JSON order (A→Z path order)
    for sec in sections:
        gmap = f"https://www.google.com/maps?q={sec['lat']},{sec['lng']}"
        # Equipment Location & Equipment
        add_row(["Equipment Location", sec["eq_location"], "", "", sec["eq_location"], sec["eq_type"], gmap], FILL_DARK_GREEN)
        add_row(["Equipment",          sec["eq_type"],     "", "", sec["eq_location"], sec["eq_type"], gmap], FILL_DARK_GREEN)

        # Cables
        for ca in sec["cables"]:
            add_row(["Cable[attach]", ca, "", "", sec["eq_location"], sec["eq_type"], ""], FILL_LIGHT_GREEN)

        # Existing Splice(s)
        for desc, cnt in sec["existing"]:
            add_row(["Existing Splice", desc, cnt, "", sec["eq_location"], sec["eq_type"], ""], FILL_YELLOW)

        # Actions for this site
        for act in site_to_actions.get(sec["site_id"], []):
            aptA = sec["pmid_to_apt"].get(act["pmid_a"], "")
            aptB = sec["pmid_to_apt"].get(act["pmid_b"], "")
            desc = (
                f"PMID: {act['pmid_a']}, Aptum ID: {aptA} .F[{act['a1']}-{act['a2']}] -- "
                f"{'Splice' if act['atype']=='Splice Required' else 'Break'} -- "
                f"PMID: {act['pmid_b']}, Aptum ID: {aptB} .F[{act['b1']}-{act['b2']}]"
            )
            add_row(
                [act["atype"], desc, act["count"], act["action_ref"], sec["eq_location"], sec["eq_type"], ""],
                FILL_ORANGE if act["atype"] == "Splice Required" else FILL_RED,
            )
            placed += 1

    # Any actions that couldn't find a site get counted here (optional: synthesize a site)
    for act in actions:
        if act["site_id"] is None:
            unplaced += 1

    # column widths, freeze header
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 35 if col == 2 else 16
    ws.freeze_panes = "A2"

    wb.save(out_path)
    return {"placed": placed, "unplaced": unplaced, "total_actions": len(actions)}


# --------------------------------------- API -------------------------------------------
def build_fiber_trace(
    json_src: Union[str, IO[str]],
    csv_src: Union[str, IO[str]],
    out_path: str = "fiber_trace_output.xlsx",
) -> Tuple[str, Dict[str, Any]]:
    """
    Build the Fiber Trace workbook.
    Returns (path_to_xlsx, stats_dict).
    """
    sections = _build_sections(json_src)
    pmid_to_sites, eq_to_site = _build_indexes(sections)
    actions = _parse_actions(csv_src, pmid_to_sites, eq_to_site)
    stats = _write_workbook(sections, actions, out_path)
    return out_path, stats
