# helper.py
from __future__ import annotations
import io, re, math
from typing import Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="AAAAAA")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def transform_fibre_action_summary_grid(wo_df: pd.DataFrame, meta: Dict) -> pd.DataFrame:
    """
    Produce a 4-column grid:
        [Label_L, Value_L, Label_R, Value_R]
    with rows matching your trace_action.csv layout.
    """
    if wo_df is None or wo_df.empty:
        wo_df = pd.DataFrame()

    # Derive columns we need
    desc_col  = _try_col(wo_df, ["Description","DESC","Desc","description"]) or (wo_df.columns[0] if len(wo_df.columns) else None)
    len_col   = _try_col(wo_df, ["Length", "End to End Length(m)", "End_to_End_Length_m"])
    otdr_col  = _try_col(wo_df, ["~OTDR Length","OTDR Length","~OTDR Length(m)","otdr_length"])

    # Counts
    num_breaks, num_splices = (0, 0)
    if desc_col:
        try:
            num_breaks, num_splices = _count_breaks_splices(wo_df, desc_col)
        except Exception:
            pass

    # Lengths
    end_to_end_len = None
    otdr_len = None
    if len_col and len_col in wo_df.columns:
        vals = pd.to_numeric(wo_df[len_col], errors="coerce").dropna()
        if len(vals):
            end_to_end_len = int(vals.sum()) if len(vals) > 1 else int(vals.iloc[0])
    if otdr_col and otdr_col in wo_df.columns:
        vals = pd.to_numeric(wo_df[otdr_col], errors="coerce").dropna()
        if len(vals):
            otdr_len = int(vals.sum()) if len(vals) > 1 else int(vals.iloc[0])

    # Pull meta (fallback to blanks)
    order_id = meta.get("order_id", "")
    wo_id    = meta.get("wo_id", "")
    designer = meta.get("designer_name", "")
    phone    = meta.get("designer_phone", "")
    date     = meta.get("date", "")
    a_end    = meta.get("a_end", "")
    z_end    = meta.get("z_end", "")
    details  = meta.get("details", "OSP DF")

    grid = [
        ["Order Number:",            order_id,                 "Number of Fibre Breaks:",   num_breaks or 0],
        ["Work Order Number:",       wo_id,                    "Number of Fibre Splices",   num_splices or 0],
        ["Order A to Z:",            f"_{a_end}_{z_end}" if a_end and z_end else "",  "End to End Length(m)",   end_to_end_len or 0],
        ["Designer:",                designer,                 "End to End ~ OTDR(m)",      otdr_len or 0],
        ["Contact Number:",          phone,                    "A END:",                    a_end],
        ["Date (dd/mm/yyyy):",       date,                     "Z END:",                    z_end],
        ["Details:",                 details,                  None,                        None],
        ["ORDER Number:",            order_id,                 None,                        None],
    ]

    # pad to a fixed height (like your sample which had blank rows beneath)
    while len(grid) < 21:
        grid.append([None, None, None, None])

    out = pd.DataFrame(grid, columns=["Order Number: ", "ORDER-267175", "Number of Fibre Breaks: ", "6"])
    return out

def read_actions_from_wo_file(uploaded_file) -> pd.DataFrame:
    raw = pd.read_csv(uploaded_file, header=None, dtype=str)
    raw.columns = range(raw.shape[1])

    hdr_idx = raw.index[raw[0].astype(str).str.strip().eq("Action")]
    if len(hdr_idx) == 0:
        return pd.DataFrame(columns=["Action", "Description", "SAP"])
    start = int(hdr_idx[0]) + 1

    sub = raw.loc[start:, [0, 1, 2]].copy()
    sub.columns = ["Action", "Description", "SAP"]
    stop_mask = sub.isna().all(axis=1)
    if stop_mask.any():
        sub = sub.loc[: stop_mask.idxmax() - 1]

    sub = sub.fillna("")
    return sub
    t = _strip_markup(s)
    # split on common separators and dedupe while keeping order
    parts = re.split(r"\s[-:]\s| ; | , ", t)
    parts = [p.strip() for p in parts if p and p.strip() != "."]
    seen, uniq = set(), []
    for p in parts:
        k = p.lower()
        if k not in seen:
            seen.add(k)
            uniq.append(p)

    # shorten verbose phrases frequently seen in source
    short = []
    for p in uniq:
        p2 = p
        p2 = re.sub(r"\bBeanfield Manhole/Handwell\b", "BF Manhole/Handwell", p2, flags=re.I)
        p2 = re.sub(r"\bUtility Pole (Wood|Concrete) With Splice Box\b",
                    r"Utility Pole (\1) + Splice Box", p2, flags=re.I)
        p2 = re.sub(r"\bHydro Manhole with Splice Box\b", "Hydro MH + Splice Box", p2, flags=re.I)
        p2 = re.sub(r"\bGeneric OSP Splice Box\b", "OSP Splice Box", p2, flags=re.I)
        p2 = re.sub(r"\bAddress[: ]", "", p2, flags=re.I)
        p2 = re.sub(r"\bToronto[, ]*", "", p2, flags=re.I)
        p2 = re.sub(r"\([^)]*\)", "", p2).strip()  # drop coords / parentheses
        short.append(p2)

    out = " | ".join([p for p in short if p])
    return re.sub(r"\s{2,}", " ", out).strip(" |")

def add_simplified_description(df: pd.DataFrame) -> pd.DataFrame:
    # pick the description column if named differently
    for cand in ["Description", "description", "DESC", "Desc", "Notes", "details", "Details"]:
        if cand in df.columns:
            desc_col = cand
            break
    else:
        desc_col = "Description"
        if desc_col not in df.columns:
            df[desc_col] = ""

    df["Description"] = df[desc_col]  # normalize name for preview
    df["Description_simplified"] = df[desc_col].apply(simplify_description)
    return df

def read_uploaded_table(up_file) -> pd.DataFrame:
    """Robust CSV reader used by other generators in your app.
    Accepts a file-like object from st.file_uploader.
    """
    up_file.seek(0)
    content = up_file.read()
    # Try UTF-8, then latin-1 as fallback
    for enc in ("utf-8", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(content), encoding=enc)
        except Exception:
            continue
    # Last resort: excel sniff
    up_file.seek(0)
    return pd.read_excel(io.BytesIO(content))

# Replace your simplify_description with this improved version:
def simplify_description(text: str) -> str:
    if pd.isna(text):
        return ""
    s = str(text).strip()
    if not s:
        return ""

    # Remove GPS coords, parentheses with only coords or IDs
    s = re.sub(r"\([^)]*\)", "", s)

    # Drop common noisy labels
    s = re.sub(r"\b(Toronto|Address|PMID|Aptum ID)\s*:?\s*", "", s, flags=re.I)

    # Replace custom tokens
    s = (s.replace("<COMMA>", ", ")
          .replace("<COLON>", ": ")
          .replace("<AND>", " & ")
          .replace("<OPEN>", "")
          .replace("<CLOSE>", ""))

    # Normalize punctuation/spacing
    s = re.sub(r"\s*[,;|]\s*", ", ", s)
    s = re.sub(r"\s{2,}", " ", s)
    s = s.strip(" ,;-")

    return s

_RX_PAIR = re.compile(
    r"^(?P<kind>Remove\s+splicing|Splice)\s+"
    r"(?P<a_id>\d+)\s*\[\s*(?P<a1>\d+)\s*[-–]\s*(?P<a2>\d+)\s*\]\s*[-–]\s*"
    r"(?P<b_id>\d+)\s*\[\s*(?P<b1>\d+)\s*[-–]\s*(?P<b2>\d+)\s*\]"
    r".*$",
    re.IGNORECASE,
)

def _fmt_span(x1: str, x2: str) -> str:
    return f"[{int(x1)}-{int(x2)}]"

def normalize_description_to_pair(text: str) -> str:
    """
    Convert verbose 'Remove splicing ... - ...' / 'Splice ... - ...' lines into the
    compact 'BREAK/Splice PMIDA [a-b] PMIDB [a-b]' strings.
    """
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    s = str(text).strip()
    if not s:
        return ""

    m = _RX_PAIR.match(s.replace("–", "-"))
    if not m:
        # No strict match; leave as-is (or return "")
        return s if s.lower() != "none" else ""

    kind = m.group("kind").lower().strip()
    a_id, a1, a2 = m.group("a_id"), m.group("a1"), m.group("a2")
    b_id, b1, b2 = m.group("b_id"), m.group("b1"), m.group("b2")

    if kind.startswith("remove"):
        # BREAK and reverse order: B then A
        return f"BREAK {b_id} {_fmt_span(b1,b2)} {a_id} {_fmt_span(a1,a2)}"
    else:
        # Splice, keep order: A then B
        return f"Splice {a_id} {_fmt_span(a1,a2)} {b_id} {_fmt_span(b1,b2)}"

def transform_fibre_action_actions(wo_df: pd.DataFrame) -> pd.DataFrame:

    """
    Extract the Action/Description/SAP section and normalize Description.
    (Uses the header-finder you already added.)
    """
    # treat wo_df as raw grid; find the real header row where col0='Action' & col1='Description'
    raw = pd.DataFrame(wo_df.values)
    raw.columns = range(raw.shape[1])

    hdr_idx = raw.index[
        raw.get(0).astype(str).str.strip().eq("Action") &
        raw.get(1).astype(str).str.strip().eq("Description")
    ]
    if len(hdr_idx) == 0:
        # case-insensitive fallback
        hdr_idx = raw.index[
            raw.get(0).astype(str).str.fullmatch(r"\s*Action\s*", case=False, na=False) &
            raw.get(1).astype(str).str.fullmatch(r"\s*Description\s*", case=False, na=False)
        ]
    if len(hdr_idx) == 0:
        return pd.DataFrame(columns=["Action", "Description", "SAP"])

    start = int(hdr_idx[0]) + 1
    sub = raw.loc[start:, [0, 1, 2]].copy()
    sub.columns = ["Action", "Description", "SAP"]

    # trim trailing blank block
    sub = sub.fillna("")
    mask_keep = ~(
        sub["Action"].astype(str).str.strip().eq("") &
        sub["Description"].astype(str).str.strip().eq("") &
        sub["SAP"].astype(str).str.strip().eq("")
    )
    sub = sub.loc[mask_keep]

    # normalize Description to your compact format
    sub["Description"] = sub["Description"].map(normalize_description_to_pair)

    # ensure Action numbering prefix like '1: Add' remains (or generate if missing)
    if not sub["Action"].astype(str).str.match(r"^\s*\d+:\s").any():
        sub["Action"] = [f"{i+1}: {a if str(a).strip() else 'Add'}"
                         for i, a in enumerate(sub["Action"].astype(str).str.strip())]

    if "SAP" not in sub.columns:
        sub["SAP"] = ""

    return sub[["Action", "Description", "SAP"]].reset_index(drop=True)

def _auto_widths(ws, df, min_w=8, max_w=60, pad=2):
    # Auto size each column using header + longest cell length
    for col_idx, col_name in enumerate(df.columns):
        header_w = len(str(col_name))
        data_w = 0 if df.empty else int(df[col_name].astype(str).map(len).max())
        width = max(header_w, data_w) + pad
        ws.set_column(col_idx, col_idx, max(min_w, min(width, max_w)))

def fibre_action_excel_bytes(summary_df: pd.DataFrame,
                             actions_df: pd.DataFrame,
                             title: str = "fibre_action") -> bytes:
    """
    Two sheets ('Summary', 'Fibre Action'), no styling, only auto-width columns.
    """
    bio = io.BytesIO()
    # NOTE: no 'options=' here to avoid the ExcelWriter.new() error
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        actions_df.to_excel(writer, sheet_name="Fibre Action", index=False)

        ws1 = writer.sheets["Summary"]
        ws2 = writer.sheets["Fibre Action"]
        _auto_widths(ws1, summary_df)
        _auto_widths(ws2, actions_df)

        # (Optional) title metadata—safe to ignore if unsupported
        try:
            writer.book.set_properties({"title": title})
        except Exception:
            pass

    bio.seek(0)
    return bio.getvalue()
    
def transform_fibre_action_summary_grid(wo_df: pd.DataFrame, meta: dict) -> pd.DataFrame:
    """
    Return a 2-column key/value DataFrame for the Summary tab.
    Overwrites the old 4-column layout.
    """
    # Derive counts if possible
    desc_col = None
    for c in wo_df.columns:
        if "desc" in c.lower():
            desc_col = c
            break
    breaks = splices = 0
    if desc_col:
        s = wo_df[desc_col].astype(str)
        breaks = s.str.contains("Remove", case=False, na=False).sum()
        splices = s.str.contains("Splice", case=False, na=False).sum()

    # Meta values
    order_id = meta.get("order_id", "")
    wo_id    = meta.get("wo_id", "")
    designer = meta.get("designer_name", "")
    phone    = meta.get("designer_phone", "")
    date     = meta.get("date", "")
    a_end    = meta.get("a_end", "")
    z_end    = meta.get("z_end", "")
    details  = meta.get("details", "OSP DF")

    # Assemble rows in desired order
    rows = [
        ("Order Number:", order_id),
        ("Work Order Number:", wo_id),
        ("Order A to Z:", f"_{a_end}_{z_end}" if a_end and z_end else ""),
        ("Designer:", designer),
        ("Contact Number:", phone),
        ("Date (dd/mm/yyyy):", date),
        ("Details:", details),
        ("Number of Fibre Breaks", breaks),
        ("Number of Fibre Splices", splices),
        ("End to End Length(m)", 0),
        ("End to End ~ OTDR(m)", 0),
        ("A END:", a_end),
        ("Z END:", z_end),
    ]

    return pd.DataFrame(rows, columns=["Field", "Value"])
