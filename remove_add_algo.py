# helper.py
from __future__ import annotations
import io, re, math
from typing import Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------- small utilities ----------
def _try_col(df: pd.DataFrame, names) -> str | None:
    for n in names:
        if n in df.columns:
            return n
    return None

def _count_breaks_splices(df: pd.DataFrame, desc_col: str) -> Tuple[int, int]:
    s = df[desc_col].astype(str).str.strip()
    num_breaks  = int(s.str.startswith(("BREAK", "Break")).sum())
    num_splices = int(s.str.startswith(("Splice", "splice")).sum())
    return num_breaks, num_splices


# ---------- 1) Summary grid (like trace_action.csv) ----------
def transform_fibre_action_summary_grid(wo_df: pd.DataFrame, meta: Dict) -> pd.DataFrame:
    """
    Produce a 4-column grid:
        [Label_L, Value_L, Label_R, Value_R]
    with rows matching your trace_action.csv layout.
    """
    if wo_df is None or wo_df.empty:
        wo_df = pd.DataFrame()

    # Derive columns we need
    desc_col  = _try_col(wo_df, ["Description","DESC","Desc","description"]) or (wo_df.columns[0] if len(wo_df.columns) else None)
    len_col   = _try_col(wo_df, ["Length", "End to End Length(m)", "End_to_End_Length_m"])
    otdr_col  = _try_col(wo_df, ["~OTDR Length","OTDR Length","~OTDR Length(m)","otdr_length"])

    # Counts
    num_breaks, num_splices = (0, 0)
    if desc_col:
        try:
            num_breaks, num_splices = _count_breaks_splices(wo_df, desc_col)
        except Exception:
            pass

    # Lengths
    end_to_end_len = None
    otdr_len = None
    if len_col and len_col in wo_df.columns:
        vals = pd.to_numeric(wo_df[len_col], errors="coerce").dropna()
        if len(vals):
            end_to_end_len = int(vals.sum()) if len(vals) > 1 else int(vals.iloc[0])
    if otdr_col and otdr_col in wo_df.columns:
        vals = pd.to_numeric(wo_df[otdr_col], errors="coerce").dropna()
        if len(vals):
            otdr_len = int(vals.sum()) if len(vals) > 1 else int(vals.iloc[0])

    # Pull meta (fallback to blanks)
    order_id = meta.get("order_id", "")
    wo_id    = meta.get("wo_id", "")
    designer = meta.get("designer_name", "")
    phone    = meta.get("designer_phone", "")
    date     = meta.get("date", "")
    a_end    = meta.get("a_end", "")
    z_end    = meta.get("z_end", "")
    details  = meta.get("details", "OSP DF")

    grid = [
        ["Order Number:",            order_id,                 "Number of Fibre Breaks:",   num_breaks or 0],
        ["Work Order Number:",       wo_id,                    "Number of Fibre Splices",   num_splices or 0],
        ["Order A to Z:",            f"_{a_end}_{z_end}" if a_end and z_end else "",  "End to End Length(m)",   end_to_end_len or 0],
        ["Designer:",                designer,                 "End to End ~ OTDR(m)",      otdr_len or 0],
        ["Contact Number:",          phone,                    "A END:",                    a_end],
        ["Date (dd/mm/yyyy):",       date,                     "Z END:",                    z_end],
        ["Details:",                 details,                  None,                        None],
        ["ORDER Number:",            order_id,                 None,                        None],
    ]

    # pad to a fixed height (like your sample which had blank rows beneath)
    while len(grid) < 21:
        grid.append([None, None, None, None])

    out = pd.DataFrame(grid, columns=["Order Number: ", "ORDER-267175", "Number of Fibre Breaks: ", "6"])
    return out


# ---------- 2) Fibre Action table (Action + Description [+ SAP blank]) ----------
def transform_fibre_action_actions(wo_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the Action/Description table (sorted by numeric prefix of Action).
    If your CSV uses different names, add them to the candidate lists below.
    """
    action_col = _try_col(wo_df, ["Action", "WO Action#", "Wo Action#", "ACTION"])
    desc_col   = _try_col(wo_df, ["Description","DESC","Desc","description"])

    if not action_col and len(wo_df.columns):
        action_col = wo_df.columns[0]
    if not desc_col and len(wo_df.columns) > 1:
        # pick a reasonable description-like column
        for c in wo_df.columns[1:]:
            if re.search(r"desc|name|device|work|scope|task", c, re.I):
                desc_col = c; break
        if not desc_col:
            desc_col = wo_df.columns[1]

    df = wo_df[[c for c in [action_col, desc_col] if c in wo_df.columns]].copy()
    df.columns = ["Action", "Description"]

    # numeric sort by the leading number in Action
    def sort_key(x):
        m = re.match(r"^\s*(\d+)", str(x))
        return int(m.group(1)) if m else 10**9
    df = df.sort_values(by="Action", key=lambda s: s.map(sort_key), kind="stable").reset_index(drop=True)

    # Add SAP col (blank) like your sheet
    df["SAP"] = ""
    return df[["Action","Description","SAP"]]


# ---------- 3) Write styled Excel ----------
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="AAAAAA")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

def _auto_width(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            widths[i] = max(widths.get(i, 10), len("" if v is None else str(v)))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(12, w + 2))

def fibre_action_excel_bytes(summary_df: pd.DataFrame, actions_df: pd.DataFrame, title="fibre_action") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # --- Summary header row (4 columns) ---
    ws.append(list(summary_df.columns))
    for j in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=j)
        c.font = BOLD; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX

    # --- Summary body ---
    for r in summary_df.itertuples(index=False):
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = LEFT
            c.border = BOX

    # widths
    _auto_width(ws)
    ws.freeze_panes = "A2"

    # --- Fibre Action sheet ---
    ws2 = wb.create_sheet("Fibre Action")
    ws2.append(list(actions_df.columns))
    for j in range(1, ws2.max_column + 1):
        c = ws2.cell(row=1, column=j)
        c.font = BOLD; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX
    for r in actions_df.itertuples(index=False):
        ws2.append(list(r))
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for c in row:
            c.alignment = LEFT
            c.border = BOX
    # common widths + freeze header
    for col, width in [("A",18), ("B",120), ("C",12)]:
        ws2.column_dimensions[col].width = width
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def read_actions_from_wo_file(uploaded_file) -> pd.DataFrame:
    raw = pd.read_csv(uploaded_file, header=None, dtype=str)
    raw.columns = range(raw.shape[1])

    hdr_idx = raw.index[raw[0].astype(str).str.strip().eq("Action")]
    if len(hdr_idx) == 0:
        return pd.DataFrame(columns=["Action", "Description", "SAP"])
    start = int(hdr_idx[0]) + 1

    sub = raw.loc[start:, [0, 1, 2]].copy()
    sub.columns = ["Action", "Description", "SAP"]
    stop_mask = sub.isna().all(axis=1)
    if stop_mask.any():
        sub = sub.loc[: stop_mask.idxmax() - 1]

    sub = sub.fillna("")
    return sub

def _strip_markup(s: str) -> str:
    if pd.isna(s):
        return ""
    t = str(s)
    repl = {
        "<COMMA>": ", ",
        "<COLON>": ": ",
        "<AND>": " & ",
        "<OPEN>": "",
        "<CLOSE>": "",
    }
    for k, v in repl.items():
        t = t.replace(k, v)
    # spacing + punctuation cleanup
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"\s*([,:;|/-])\s*", r" \1 ", t)
    t = re.sub(r"\s{2,}", " ", t)
    # canonicalize tokens
    t = re.sub(r"\bPMID\s*[:\-]?\s*(\d+)", r"PMID \1", t, flags=re.I)
    t = re.sub(r"\bAptum ID\s*[:\-]?\s*([.\w-]+)", r"Aptum ID \1", t, flags=re.I)
    return t.strip(" -|;,. ")

def simplify_description(s: str) -> str:
    t = _strip_markup(s)
    # split on common separators and dedupe while keeping order
    parts = re.split(r"\s[-:]\s| ; | , ", t)
    parts = [p.strip() for p in parts if p and p.strip() != "."]
    seen, uniq = set(), []
    for p in parts:
        k = p.lower()
        if k not in seen:
            seen.add(k)
            uniq.append(p)

    # shorten verbose phrases frequently seen in source
    short = []
    for p in uniq:
        p2 = p
        p2 = re.sub(r"\bBeanfield Manhole/Handwell\b", "BF Manhole/Handwell", p2, flags=re.I)
        p2 = re.sub(r"\bUtility Pole (Wood|Concrete) With Splice Box\b",
                    r"Utility Pole (\1) + Splice Box", p2, flags=re.I)
        p2 = re.sub(r"\bHydro Manhole with Splice Box\b", "Hydro MH + Splice Box", p2, flags=re.I)
        p2 = re.sub(r"\bGeneric OSP Splice Box\b", "OSP Splice Box", p2, flags=re.I)
        p2 = re.sub(r"\bAddress[: ]", "", p2, flags=re.I)
        p2 = re.sub(r"\bToronto[, ]*", "", p2, flags=re.I)
        p2 = re.sub(r"\([^)]*\)", "", p2).strip()  # drop coords / parentheses
        short.append(p2)

    out = " | ".join([p for p in short if p])
    return re.sub(r"\s{2,}", " ", out).strip(" |")

def add_simplified_description(df: pd.DataFrame) -> pd.DataFrame:
    # pick the description column if named differently
    for cand in ["Description", "description", "DESC", "Desc", "Notes", "details", "Details"]:
        if cand in df.columns:
            desc_col = cand
            break
    else:
        desc_col = "Description"
        if desc_col not in df.columns:
            df[desc_col] = ""

    df["Description"] = df[desc_col]  # normalize name for preview
    df["Description_simplified"] = df[desc_col].apply(simplify_description)
    return df

    
def read_uploaded_table(up_file) -> pd.DataFrame:
    """Robust CSV reader used by other generators in your app.
    Accepts a file-like object from st.file_uploader.
    """
    up_file.seek(0)
    content = up_file.read()
    # Try UTF-8, then latin-1 as fallback
    for enc in ("utf-8", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(content), encoding=enc)
        except Exception:
            continue
    # Last resort: excel sniff
    up_file.seek(0)
    return pd.read_excel(io.BytesIO(content))

# ---------- Summary grid ----------

def transform_fibre_action_summary_grid(wo_df: pd.DataFrame, meta: dict) -> pd.DataFrame:
    """Return a 2xN key/value grid flattened to 4 columns: L1,V1,L2,V2.
    Matches the style of trace_action.csv Summary sheet.
    """
    rows = [
        ("Order #", meta.get("order_id", "")),
        ("WO #", meta.get("wo_id", "")),
        ("Designer", meta.get("designer_name", "")),
        ("Phone", meta.get("designer_phone", "")),
        ("Date", meta.get("date", "")),
        ("A-End", meta.get("a_end", "")),
        ("Z-End", meta.get("z_end", "")),
        ("Details", meta.get("details", "")),
    ]
    # pack into 2 columns per row, 2 rows side-by-side -> 4 columns
    left = rows[0:len(rows)//2 + len(rows)%2]
    right= rows[len(rows)//2 + len(rows)%2 :]
    # pad right side if needed
    while len(right) < len(left):
        right.append(("", ""))
    data = []
    for (k1,v1),(k2,v2) in zip(left, right):
        data.append({"L1":k1, "V1":v1, "L2":k2, "V2":v2})
    return pd.DataFrame(data)

# ---------- Actions table ----------

def transform_fibre_action_actions(wo_df: pd.DataFrame) -> pd.DataFrame:
    """Derive an Action/Description/SAP table from the WO CSV.
    This keeps your existing logic lightweight and predictable:
    - If the CSV already has these columns, use them directly.
    - Else, look for any columns that look like 'Action' or 'Description'.
    - Else, create an empty shell so the UI stays stable.
    """
    cols = {c.lower(): c for c in wo_df.columns}

    if {"action","description"}.issubset(cols.keys()):
        df = wo_df[[cols["action"], cols["description"]]].copy()
        df.columns = ["Action","Description"]
    elif "action" in cols:
        df = wo_df[[cols["action"]]].copy()
        df["Description"] = ""
    else:
        # Minimal fallback – enumerate by row index
        df = pd.DataFrame({
            "Action": [f"{i+1}: Add" for i in range(len(wo_df))],
            "Description": ["" for _ in range(len(wo_df))],
        })

    # Ensure numeric prefixes like "1: "
    if not df["Action"].str.contains(r"^\d+:\s").any():
        df["Action"] = [f"{i+1}: {a}" for i, a in enumerate(df["Action"].astype(str).str.replace(r"^\d+:\s*", "", regex=True))]

    # Always include SAP for alignment
    df["SAP"] = ""
    return df[["Action","Description","SAP"]]

# ---------- Description simplifier ----------

def simplify_description(text: str) -> str:
    """Make descriptions short and human-friendly like the reference sheet.
    Rules (in order):
      1) If empty/NaN -> "None".
      2) Remove GPS coords, parentheses-only fragments, and long numeric IDs.
      3) Drop noisy prefixes like 'Toronto,', 'Address:', 'PMID:', 'Aptum ID:'.
      4) Collapse multiple separators to a single comma and space.
      5) Trim to ~120 chars (without hard-cutting words) and strip.
    """
    if pd.isna(text) or str(text).strip() == "":
        return "None"
    s = str(text)
    # Remove lat/long and parentheses blocks with only symbols/numbers/commas
    s = re.sub(r"\(.*?\)", "", s)
    s = re.sub(r"\b-?\d{2,}\.?\d*\s*,\s*-?\d{2,}\.?\d*\b", "", s)  # coords
    # Drop noisy labels
    s = re.sub(r"\b(Toronto|Address|PMID|Aptum ID)\s*:?\s*", "", s, flags=re.IGNORECASE)
    # Replace special tokens
    s = s.replace("<COMMA>", ", ").replace("<AND>", " & ")
    s = s.replace("<OPEN>", "").replace("<CLOSE>", "")
    # Smash repeated punctuation/spaces
    s = re.sub(r"[;|]\s*", ", ", s)
    s = re.sub(r"\s*[:,]\s*", ", ", s)
    s = re.sub(r"\s{2,}", " ", s)
    s = re.sub(r",\s*,+", ", ", s)
    s = s.strip(" ,-")
    # Friendly cap at ~120 chars
    if len(s) > 120:
        cut = s[:120].rsplit(" ", 1)[0]
        s = f"{cut}…"
    return s if s else "None"

# ---------- Excel writer ----------

def fibre_action_excel_bytes(summary_df: pd.DataFrame, actions_df: pd.DataFrame, title: str = "fibre_action") -> bytes:
    """Create a styled two-tab Excel file in-memory.
    Tabs: 'Summary' and 'Fibre Action'.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as xw:
        # --- Summary sheet ---
        summary_df.to_excel(xw, sheet_name="Summary", index=False, header=["Action","Description","SAP"] if list(summary_df.columns)==[0,1,2] else True)
        wb  = xw.book
        ws1 = xw.sheets["Summary"]
        header_fmt = wb.add_format({"bold": True, "bg_color": "#1f2937", "font_color": "#ffffff", "border": 1, "align": "center"})
        body_fmt   = wb.add_format({"border":1})
        for col, _ in enumerate(summary_df.columns):
            ws1.set_column(col, col, 28, body_fmt)
        ws1.set_row(0, 18, header_fmt)

        # --- Fibre Action sheet ---
        fa = actions_df[["Action","Description","SAP"]].copy()
        fa.to_excel(xw, sheet_name="Fibre Action", index=False)
        ws2 = xw.sheets["Fibre Action"]
        hdr2 = wb.add_format({"bold": True, "bg_color": "#1f2937", "font_color": "#ffffff", "border": 1, "align": "center"})
        cell = wb.add_format({"border":1, "text_wrap": True})
        ws2.set_row(0, 18, hdr2)
        ws2.set_column(0, 0, 18, cell)  # Action
        ws2.set_column(1, 1, 70, cell)  # Description
        ws2.set_column(2, 2, 12, cell)  # SAP

        # Freeze header rows for both sheets
        ws1.freeze_panes(1, 0)
        ws2.freeze_panes(1, 0)

        # Optional title property
        wb.set_properties({"title": title})

    return output.getvalue()

